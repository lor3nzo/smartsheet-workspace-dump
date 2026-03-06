"""
Smartsheet Full Workspace Dump -> Excel / CSV / Parquet

Requirements:
    pip install smartsheet-python-sdk openpyxl pandas python-dotenv
    pip install pyarrow   # only required for --output-format parquet

Usage:
    python smartsheet_workspace_dump.py
    python smartsheet_workspace_dump.py --workspace-id 123 --output dump.xlsx
    python smartsheet_workspace_dump.py --values both --log-level DEBUG --dry-run
    python smartsheet_workspace_dump.py --row-metadata
    python smartsheet_workspace_dump.py --output-format csv
    python smartsheet_workspace_dump.py --output-format both
    python smartsheet_workspace_dump.py --include-regex "waverly" --max-sheets 5
    python smartsheet_workspace_dump.py --format minimal --validation-level basic
    python smartsheet_workspace_dump.py --format pretty --autofit-rows 200 --autofit-max-width 80 --validation-level deep
    python smartsheet_workspace_dump.py --no-index --no-summary --format minimal
    python smartsheet_workspace_dump.py --validation-level deep --max-validation-issues 0
    python smartsheet_workspace_dump.py --since 2026-03-01
    python smartsheet_workspace_dump.py --since last-run
"""

from dotenv import load_dotenv
load_dotenv()

import warnings
# Suppress known SDK DeprecationWarnings (include_all, get_workspace, ssl options).
# The .bat runner also passes -W ignore::DeprecationWarning as the reliable catch-all.
warnings.filterwarnings("ignore", category=DeprecationWarning)

import argparse
import hashlib
import json
import logging
import math
import os
import random
import re
import shutil
import sys
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Optional

import pandas as pd
import smartsheet
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# -- CONSTANTS ----------------------------------------------------------------
DEFAULT_OUTPUT   = "smartsheet.xlsx"
ARCHIVE_DIR      = "archive"
LOG_FILE         = "smartsheet_dump.log"
MAX_RETRIES      = 3
RETRY_BASE_DELAY = 2.0
MAX_WORKERS      = 5
INDENT_COL       = "_Indent_Level"
INDEX_TAB        = "INDEX"
ROW_META_COLS    = ["_Row_ID", "_Parent_Row_ID", "_Row_Number", "_Created_At", "_Modified_At"]
# ----------------------------------------------------------------------------


def _normalize(v) -> str:
    """Normalize cell values for write validation: NaN/None -> '', whole floats -> int, dates -> ISO."""
    if v is None:
        return ""
    if isinstance(v, float):
        if math.isnan(v):
            return ""
        if v == int(v):
            return str(int(v))
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


# -- CLI ----------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Dump Smartsheet workspaces to Excel")
    p.add_argument("--workspace-id", default=os.environ.get("SMARTSHEET_WORKSPACE_ID"),
                   help="Single workspace ID (default: all workspaces)")
    p.add_argument("--output",    default=DEFAULT_OUTPUT,
                   help="Output .xlsx filename")
    p.add_argument("--values",    choices=["raw", "display", "both"], default="both",
                   help="Cell value export mode (default: both)")
    p.add_argument("--log-level", choices=["DEBUG", "INFO", "WARNING", "ERROR"], default="INFO")
    p.add_argument("--dry-run",   action="store_true",
                   help="Discover and list sheets without writing any file")
    p.add_argument("--log-every", type=int, default=10,
                   help="Log progress every N sheets (default: 10)")
    p.add_argument("--workers",       type=int, default=MAX_WORKERS,
                   help=f"Parallel fetch workers (default: {MAX_WORKERS})")
    p.add_argument("--include-regex", default=None,
                   help="Only export sheets whose name matches this regex (case-insensitive)")
    p.add_argument("--exclude-regex", default=None,
                   help="Skip sheets whose name matches this regex (case-insensitive)")
    p.add_argument("--max-sheets",    type=int, default=None,
                   help="Cap total sheets exported (useful for testing)")
    p.add_argument("--row-metadata",  action="store_true", default=False,
                   help="Append _Row_ID, _Parent_Row_ID, _Row_Number, _Created_At, _Modified_At to each sheet")
    p.add_argument("--output-format", choices=["xlsx", "csv", "parquet", "both", "all"], default="xlsx",
                   help="Output format: xlsx (default), csv, parquet, both (xlsx+csv), all (xlsx+csv+parquet)")
    p.add_argument("--format",         choices=["pretty", "minimal"], default="pretty",
                   help="Workbook formatting: pretty (default, styled+autofit) or minimal (freeze panes only, faster)")
    p.add_argument("--autofit-rows",   type=int, default=50,
                   help="Max rows to sample for column autofit (default: 50, 0=disable). Ignored under --format minimal.")
    p.add_argument("--validation-level", choices=["basic", "standard", "deep"], default="standard",
                   help="basic=counts+headers, standard=+sampled cells (default), deep=all rows hashed (slow)")
    p.add_argument("--no-index",   action="store_true", default=False,
                   help="Omit the INDEX tab (faster for automated/scripted runs)")
    p.add_argument("--no-summary", action="store_true", default=False,
                   help="Omit the RUN_SUMMARY and SKIPPED tabs")
    p.add_argument("--autofit-max-width", type=int, default=60,
                   help="Maximum column width in autofit (default: 60). Ignored under --format minimal.")
    p.add_argument("--max-validation-issues", type=int, default=10,
                   help="Max issues logged per sheet in deep validation mode (default: 10, 0=unlimited)")
    p.add_argument("--since",      default=None,
                   help="Incremental export: only re-fetch sheets modified after this date (ISO: 2026-01-01) or 'last-run'")
    p.add_argument("--state-file", default=None,
                   help="Path to incremental state sidecar JSON (default: {output}.state.json)")
    return p.parse_args()


# -- LOGGING ------------------------------------------------------------------
def setup_logging(level: str) -> logging.Logger:
    logger    = logging.getLogger("ss_dump")
    log_level = getattr(logging, level, logging.INFO)
    logger.setLevel(log_level)                      # always refresh level
    if logger.handlers:
        for h in logger.handlers:                   # refresh level on existing handlers too
            h.setLevel(log_level)
        return logger
    fmt = logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s", "%Y-%m-%d %H:%M:%S")
    for h in [logging.StreamHandler(sys.stdout), logging.FileHandler(LOG_FILE, encoding="utf-8")]:
        h.setFormatter(fmt)
        h.setLevel(log_level)
        logger.addHandler(h)
    return logger


# -- DATA CLASSES -------------------------------------------------------------
@dataclass
class SheetRecord:
    orig_name:      str
    sheet_id:       int
    workspace_name: str      = ""
    modified_at:    Optional[datetime] = None

@dataclass
class FetchedSheet:
    record:   SheetRecord
    df:       pd.DataFrame
    tab_name: str

@dataclass
class ExportStats:
    total_found:       int   = 0
    exported:          int   = 0
    skipped_type:      int   = 0
    skipped_error:     int   = 0
    validation_issues: list  = field(default_factory=list)
    errors:            list  = field(default_factory=list)
    elapsed:           float = 0.0


# -- RETRY --------------------------------------------------------------------
TRANSIENT_API_CODES = frozenset([429, 502, 503])  # 502 = Bad Gateway (transient infra error)

def with_retry(fn, *args, logger, **kwargs):
    """Call fn(*args, **kwargs) with exponential backoff + jitter on transient errors."""
    for attempt in range(1, MAX_RETRIES + 2):
        try:
            return fn(*args, **kwargs)
        except smartsheet.exceptions.ApiError as e:
            code = getattr(getattr(e, "error", None), "status_code", None)
            if attempt <= MAX_RETRIES and code in TRANSIENT_API_CODES:
                delay = RETRY_BASE_DELAY * (2 ** (attempt - 1)) + random.uniform(0, 0.5)
                logger.warning(f"API {code} (attempt {attempt}/{MAX_RETRIES}), retry in {delay:.1f}s")
                time.sleep(delay)
            else:
                raise
        except TypeError as e:
            # SDK bug: malformed API response (e.g. 502 with no body) causes
            # NoneType concatenation inside smartsheet.py before ApiError is raised.
            # Treat as transient -- retry up to MAX_RETRIES times.
            if attempt <= MAX_RETRIES:
                delay = RETRY_BASE_DELAY * (2 ** (attempt - 1)) + random.uniform(0, 0.5)
                logger.warning(f"SDK TypeError on malformed response (attempt {attempt}/{MAX_RETRIES}), retry in {delay:.1f}s: {e}")
                time.sleep(delay)
            else:
                raise
        except (ConnectionError, TimeoutError, OSError,
                smartsheet.exceptions.SystemMaintenanceError) as e:
            if attempt <= MAX_RETRIES:
                delay = RETRY_BASE_DELAY * (2 ** (attempt - 1)) + random.uniform(0, 0.5)
                logger.warning(f"Network error (attempt {attempt}/{MAX_RETRIES}), retry in {delay:.1f}s: {type(e).__name__}: {e}")
                time.sleep(delay)
            else:
                raise


# -- NAMING -------------------------------------------------------------------
def sanitize_sheet_name(name: str, fallback: str = "Sheet") -> str:
    name = re.sub(r"[\\/*?\[\]:]", "", name).strip()
    return name[:31] if name else fallback[:31]


def unique_tab_name(raw_name: str, sheet_id: int, seen: dict) -> str:
    """Case-insensitive unique tab name, suffix fits within Excel's 31-char limit."""
    base = sanitize_sheet_name(raw_name, fallback=f"Sheet_{sheet_id}")
    key  = base.lower()
    if key not in seen:
        seen[key] = 0
        return base
    seen[key] += 1
    suffix   = f"_{seen[key]}"
    tab_name = base[:31 - len(suffix)] + suffix
    seen[tab_name.lower()] = 0
    return tab_name


def _safe_reserved_name(base: str, seen: dict) -> str:
    """Return a tab name that doesn't collide with real sheets."""
    name = base
    while name.lower() in seen:
        name += "_"
    return name

def safe_index_name(seen: dict) -> str:
    return _safe_reserved_name(INDEX_TAB, seen)

def safe_summary_name(seen: dict) -> str:
    return _safe_reserved_name("RUN_SUMMARY", seen)

def safe_skipped_name(seen: dict) -> str:
    return _safe_reserved_name("SKIPPED", seen)


# -- DISCOVERY ----------------------------------------------------------------
def _parse_modified_at(sheet_obj) -> Optional[datetime]:
    """
    Extract modified_at from a Smartsheet sheet object.
    The SDK may return a datetime, an ISO string, or None depending on version.
    Returns timezone-naive datetime or None on any failure.
    """
    raw = getattr(sheet_obj, "modified_at", None)
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.replace(tzinfo=None)   # strip tz for consistent naive comparison
    try:
        dt = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
        return dt.replace(tzinfo=None)
    except (ValueError, AttributeError):
        return None
def _collect_from_folder(folder, workspace_name: str) -> list:
    records = []
    if getattr(folder, "sheets", None):
        for s in folder.sheets:
            records.append(SheetRecord(s.name, s.id, workspace_name, _parse_modified_at(s)))
    if getattr(folder, "folders", None):
        for sub in folder.folders:
            records.extend(_collect_from_folder(sub, workspace_name))
    return records


def discover_sheets(client, workspace_id: Optional[str], logger) -> list:
    records = []

    if workspace_id:
        workspaces = [with_retry(
            client.Workspaces.get_workspace, int(workspace_id),
            logger=logger, include="sheets,folders"
        )]
    else:
        ws_list    = with_retry(client.Workspaces.list_workspaces, logger=logger, include_all=True)
        workspaces = [
            with_retry(client.Workspaces.get_workspace, ws.id,
                       logger=logger, include="sheets,folders")
            for ws in ws_list.data
        ]

    for ws in workspaces:
        if getattr(ws, "sheets", None):
            for s in ws.sheets:
                records.append(SheetRecord(s.name, s.id, ws.name, _parse_modified_at(s)))
        if getattr(ws, "folders", None):
            for folder in ws.folders:
                records.extend(_collect_from_folder(folder, ws.name))

    if not workspace_id:
        ws_ids = {r.sheet_id for r in records}
        home   = with_retry(client.Sheets.list_sheets, logger=logger, include_all=True)
        for s in home.data:
            if s.id not in ws_ids:
                records.append(SheetRecord(s.name, s.id, "(Home)", _parse_modified_at(s)))

    logger.info(f"Discovered {len(records)} sheet(s) across {len(workspaces)} workspace(s)")

    # Final dedup pass -- guard against duplicates across recursive paths
    seen_ids, deduped = set(), []
    for r in records:
        if r.sheet_id not in seen_ids:
            seen_ids.add(r.sheet_id)
            deduped.append(r)
    if len(deduped) < len(records):
        logger.warning(f"Removed {len(records) - len(deduped)} duplicate sheet(s) after discovery")
    return deduped


# -- EXTRACTION ---------------------------------------------------------------
def _resolve_col_titles(columns) -> dict:
    """Map col.id -> unique title; append _2, _3 for duplicate column names."""
    seen, result = {}, {}
    for col in columns:
        title = col.title or f"Col_{col.id}"
        if title in seen:
            seen[title] += 1
            title = f"{title}_{seen[title]}"
        else:
            seen[title] = 1
        result[col.id] = title
    return result


def extract_sheet(client, record: SheetRecord, value_mode: str, logger,
                  row_metadata: bool = False) -> Optional[pd.DataFrame]:
    """
    Fetch one sheet and return a DataFrame.
    Returns None if the object is not a plain sheet (report, dashboard, etc).
    Includes _Indent_Level column to preserve row hierarchy.
    Optionally appends row metadata columns when row_metadata=True.
    """
    raw      = with_retry(client.Sheets.get_sheet, record.sheet_id, logger=logger)
    obj_type = getattr(raw, "type", None)
    if obj_type and str(obj_type).upper() not in ("SHEET", ""):
        logger.warning(f"Skipping '{record.orig_name}': unsupported object type '{obj_type}'")
        return None

    col_map    = _resolve_col_titles(raw.columns)
    col_titles = list(col_map.values())

    if value_mode == "both":
        data_cols = [c for t in col_titles for c in (t, f"{t}_raw")]
    else:
        data_cols = col_titles

    all_cols = [INDENT_COL] + data_cols
    if row_metadata:
        all_cols = all_cols + ROW_META_COLS

    rows = []
    for row in raw.rows:
        row_data             = dict.fromkeys(all_cols)
        row_data[INDENT_COL] = getattr(row, "indent", 0) or 0
        for cell in row.cells:
            title = col_map.get(cell.column_id)
            if not title:
                continue
            display = cell.display_value if cell.display_value is not None else cell.value
            if value_mode == "raw":
                row_data[title] = cell.value
            elif value_mode == "display":
                row_data[title] = display
            else:
                row_data[title]          = display
                row_data[f"{title}_raw"] = cell.value

        if row_metadata:
            row_data["_Row_ID"]        = getattr(row, "id", None)
            row_data["_Parent_Row_ID"] = getattr(row, "parent_id", None)
            row_data["_Row_Number"]    = getattr(row, "row_number", None)
            created  = getattr(row, "created_at", None)
            modified = getattr(row, "modified_at", None)
            row_data["_Created_At"]  = created.isoformat()  if created  else None
            row_data["_Modified_At"] = modified.isoformat() if modified else None

        rows.append(row_data)

    return pd.DataFrame(rows, columns=all_cols)


# -- RENDERING ----------------------------------------------------------------
def style_header_row(ws, row: int = 1):
    fill  = PatternFill("solid", start_color="1F3864")
    font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[row]:
        cell.fill, cell.font, cell.alignment = fill, font, align
    ws.row_dimensions[row].height = 20


MAX_AUTOFIT_ROWS = 50   # default sample cap -- overridable via --autofit-rows

def auto_fit_columns(ws, max_rows: int = MAX_AUTOFIT_ROWS, max_width: int = 60):
    if max_rows == 0:
        return   # auto-fit disabled
    col_widths = {}
    for i, row in enumerate(ws.iter_rows()):
        if i >= max_rows:
            break
        for cell in row:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), len(str(cell.value)))
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = min(width + 4, max_width)


def build_index_sheet(wb, manifest: list, index_tab: str,
                      fmt_mode: str = "pretty", autofit_rows: int = MAX_AUTOFIT_ROWS,
                      max_width: int = 60):
    idx = wb.create_sheet(index_tab, 0)

    # Row 1: metadata (not part of the data table)
    idx.append(["", "", "", "", "", "", f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    if fmt_mode == "pretty":
        meta_font = Font(italic=True, name="Arial", size=9, color="888888")
        idx.cell(row=1, column=7).font = meta_font

    # Row 2: column headers
    idx.append(["#", "Workspace", "Sheet Name", "Rows", "Smartsheet ID", "Link", ""])
    if fmt_mode == "pretty":
        style_header_row(idx, row=2)

    for i, fs in enumerate(manifest, start=1):
        url = f"https://app.smartsheet.com/sheets/{fs.record.sheet_id}"
        idx.append([i, fs.record.workspace_name, fs.record.orig_name, len(fs.df), fs.record.sheet_id, "", ""])
        data_row = i + 2  # offset by metadata row

        name_cell = idx.cell(row=data_row, column=3)
        needs_quote = bool(re.search(r"[ '\[\]!]", fs.tab_name))
        escaped     = fs.tab_name.replace("'", "''")
        safe        = f"'{escaped}'" if needs_quote else fs.tab_name
        name_cell.hyperlink = f"#{safe}!A1"
        name_cell.font      = Font(color="0070C0", underline="single", name="Arial", size=10)

        link_cell           = idx.cell(row=data_row, column=6)
        link_cell.value     = "Open"
        link_cell.hyperlink = url
        link_cell.font      = Font(color="0070C0", underline="single", name="Arial", size=10)

    if fmt_mode == "pretty":
        auto_fit_columns(idx, max_rows=autofit_rows, max_width=max_width)


# -- SUMMARY / SKIPPED TABS ---------------------------------------------------
def build_summary_sheet(wb, stats: "ExportStats", args, summary_name: str, skipped_name: str,
                        fmt_mode: str = "pretty", autofit_rows: int = MAX_AUTOFIT_ROWS,
                        max_width: int = 60):
    """RUN_SUMMARY tab: operational evidence inside the workbook."""
    ws = wb.create_sheet(summary_name)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    rows = [
        ("Run timestamp",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Output file",      args.output),
        ("Value mode",       args.values),
        ("Format mode",      args.format),
        ("Autofit rows",     args.autofit_rows if args.format == "pretty" else "n/a (minimal)"),
        ("Autofit max width", args.autofit_max_width if args.format == "pretty" else "n/a (minimal)"),
        ("Validation level", args.validation_level),
        ("Max issues/sheet", args.max_validation_issues if args.validation_level == "deep" else "n/a"),
        ("Include index",    "no" if args.no_index else "yes"),
        ("Include summary",  "no" if args.no_summary else "yes"),
        ("Since",            args.since or "full run"),
        ("State file",       _state_path(args) if args.since else "n/a"),
        ("Workspace filter", args.workspace_id or "All"),
        ("Elapsed (s)",      f"{stats.elapsed:.1f}"),
        ("",                 ""),
        ("Sheets found",     stats.total_found),
        ("Sheets exported",  stats.exported),
        ("Skipped (type)",   stats.skipped_type),
        ("Skipped (error)",  stats.skipped_error),
        ("Validation issues", len(stats.validation_issues)),
    ]
    for r in rows:
        ws.append(r)

    if fmt_mode == "pretty":
        style_header = Font(bold=True, name="Arial", size=10)
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value:
                    cell.font = style_header

    if stats.errors or stats.validation_issues:
        skipped_ws = wb.create_sheet(skipped_name)
        skipped_ws.append(["Sheet Name", "Reason"])
        if fmt_mode == "pretty":
            style_header_row(skipped_ws)
        for msg in stats.errors:
            skipped_ws.append(["", msg])
        for msg in stats.validation_issues:
            skipped_ws.append(["", f"VALIDATION ISSUE: {msg}"])
        if fmt_mode == "pretty":
            auto_fit_columns(skipped_ws, max_rows=autofit_rows, max_width=max_width)


# -- SUMMARY LOGGER -----------------------------------------------------------
def _log_summary(stats: "ExportStats", logger, output: str = ""):
    logger.info("-" * 50)
    if output:
        logger.info(f"Output:            {output}")
    logger.info(f"Elapsed:           {stats.elapsed:.1f}s")
    logger.info(f"Found:             {stats.total_found}")
    logger.info(f"Exported:          {stats.exported}")
    logger.info(f"Skipped (type):    {stats.skipped_type}")
    logger.info(f"Skipped (error):   {stats.skipped_error}")
    if stats.validation_issues:
        logger.warning(f"Validation issues: {len(stats.validation_issues)}")
        for m in stats.validation_issues:
            logger.warning(f"  {m}")
    if stats.errors:
        logger.error("Errors:")
        for e in stats.errors:
            logger.error(f"  {e}")
    logger.info("-" * 50)


_WINDOWS_RESERVED = frozenset([
    "CON", "PRN", "AUX", "NUL",
    "COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9",
    "LPT1", "LPT2", "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9",
])
MAX_FILENAME_LEN = 100  # conservative cross-platform cap


def _safe_filename_stem(name: str, sheet_id: int) -> str:
    """Return a filesystem-safe filename stem (no extension)."""
    # Strip invalid chars
    stem = re.sub(r"[\\/*?:\[\]\"<>|]", "_", name).strip()
    # Remove trailing dots and spaces (Windows rejects these)
    stem = stem.rstrip(". ")
    # Replace Windows reserved names (case-insensitive)
    if stem.upper() in _WINDOWS_RESERVED:
        stem = f"{stem}_{sheet_id}"
    # Enforce length cap
    stem = stem[:MAX_FILENAME_LEN] if stem else f"Sheet_{sheet_id}"
    return stem or f"Sheet_{sheet_id}"


def _unique_filename(orig_name: str, sheet_id: int, seen: dict) -> str:
    """Return a unique, filesystem-safe filename stem for flat-file output."""
    base = _safe_filename_stem(orig_name, sheet_id)
    key  = base.lower()
    if key not in seen:
        seen[key] = 0
        return base
    seen[key] += 1
    return f"{base}_{seen[key]}"




# -- FLAT-FILE WRITERS --------------------------------------------------------
def write_csv_output(manifest: list, stem: str, logger, run_ts: str = "") -> str:
    """Write one CSV per sheet into {stem}_csv/{run_ts}/. Copies to _latest/. Returns run dir."""
    if not run_ts:
        run_ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    base_dir = f"{stem}_csv"
    run_dir  = os.path.join(base_dir, run_ts)
    os.makedirs(run_dir, exist_ok=True)

    seen_names    = {}
    manifest_rows = []
    for fs in manifest:
        filename = _unique_filename(fs.record.orig_name, fs.record.sheet_id, seen_names)
        path     = os.path.join(run_dir, f"{filename}.csv")
        fs.df.to_csv(path, index=False, encoding="utf-8-sig")
        manifest_rows.append({
            "sheet_name": fs.record.orig_name,
            "workspace":  fs.record.workspace_name,
            "sheet_id":   fs.record.sheet_id,
            "rows":       len(fs.df),
            "file":       f"{filename}.csv",
        })

    pd.DataFrame(manifest_rows).to_csv(
        os.path.join(run_dir, "_manifest.csv"), index=False, encoding="utf-8-sig"
    )

    # _latest/ = stable pointer to most recent run
    latest_dir = os.path.join(base_dir, "_latest")
    if os.path.isdir(latest_dir):
        shutil.rmtree(latest_dir)
    shutil.copytree(run_dir, latest_dir)

    logger.info(f"CSV output written -> {run_dir}/ ({len(manifest)} file(s) + _manifest.csv)")
    logger.info(f"  _latest/ updated -> {latest_dir}/")
    return run_dir


def write_parquet_output(manifest: list, stem: str, logger, run_ts: str = "") -> str:
    """Write one Parquet per sheet into {stem}_parquet/{run_ts}/. Copies to _latest/. Returns run dir."""
    try:
        import pyarrow  # noqa: F401 -- presence check only
    except ImportError:
        logger.error("parquet output requires pyarrow. Install it with: pip install pyarrow")
        sys.exit(1)

    if not run_ts:
        run_ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    base_dir = f"{stem}_parquet"
    run_dir  = os.path.join(base_dir, run_ts)
    os.makedirs(run_dir, exist_ok=True)

    seen_names    = {}
    manifest_rows = []
    for fs in manifest:
        filename = _unique_filename(fs.record.orig_name, fs.record.sheet_id, seen_names)
        path     = os.path.join(run_dir, f"{filename}.parquet")
        fs.df.to_parquet(path, index=False)
        manifest_rows.append({
            "sheet_name": fs.record.orig_name,
            "workspace":  fs.record.workspace_name,
            "sheet_id":   fs.record.sheet_id,
            "rows":       len(fs.df),
            "file":       f"{filename}.parquet",
        })

    pd.DataFrame(manifest_rows).to_csv(
        os.path.join(run_dir, "_manifest.csv"), index=False, encoding="utf-8-sig"
    )

    # _latest/ = stable pointer to most recent run
    latest_dir = os.path.join(base_dir, "_latest")
    if os.path.isdir(latest_dir):
        shutil.rmtree(latest_dir)
    shutil.copytree(run_dir, latest_dir)

    logger.info(f"Parquet output written -> {run_dir}/ ({len(manifest)} file(s) + _manifest.csv)")
    logger.info(f"  _latest/ updated -> {latest_dir}/")
    return run_dir


def _check_parquet_early(fmt: str, logger):
    """Fail fast before any API calls if parquet is requested but pyarrow is missing."""
    if fmt in ("parquet", "all"):
        try:
            import pyarrow  # noqa: F401
        except ImportError:
            logger.error("parquet output requires pyarrow. Install it with: pip install pyarrow")
            sys.exit(1)


# -- INCREMENTAL STATE --------------------------------------------------------
def _state_path(args) -> str:
    """Resolve sidecar path: explicit --state-file or {output}.state.json."""
    return args.state_file or f"{args.output}.state.json"


def _load_state(path: str, logger) -> dict:
    """Load sidecar JSON. Returns empty dict on missing/corrupt file."""
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        logger.warning(f"Could not read state file {path}: {e}. Treating as full run.")
        return {}


def _save_state(path: str, state: dict, logger) -> None:
    """Persist sidecar JSON atomically."""
    tmp = f"{path}.tmp_{uuid.uuid4().hex[:8]}"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2)
        os.replace(tmp, path)
    except OSError as e:
        logger.warning(f"Could not save state file {path}: {e}")


def _parse_since(since_str: str, state: dict, output_key: str, logger):
    """
    Parse --since value into a timezone-naive datetime for comparison.
    Returns None (full run) if unparseable or no prior state.
    """
    if since_str == "last-run":
        last = state.get(output_key, {}).get("last_run")
        if not last:
            logger.info("--since last-run: no prior state found, running full export.")
            return None
        try:
            return datetime.fromisoformat(last)
        except ValueError:
            logger.warning(f"--since last-run: corrupt timestamp '{last}', running full export.")
            return None
    try:
        return datetime.fromisoformat(since_str)
    except ValueError:
        logger.error(f"--since: invalid date '{since_str}'. Use ISO format (2026-01-01) or 'last-run'.")
        sys.exit(1)


def _read_sheet_from_xlsx(xlsx_path: str, tab_name: str, logger) -> "Optional[pd.DataFrame]":
    """Read one tab from an existing XLSX back into a DataFrame (cache hit for incremental mode)."""
    try:
        return pd.read_excel(xlsx_path, sheet_name=tab_name, engine="openpyxl")
    except Exception as e:
        logger.warning(f"Could not read cached tab '{tab_name}' from {xlsx_path}: {e}. Will re-fetch.")
        return None


# -- MAIN ---------------------------------------------------------------------
def main():
    args   = parse_args()
    logger = setup_logging(args.log_level)
    log_n  = max(1, args.log_every)

    # -- Argument validation --------------------------------------------------
    if args.autofit_rows < 0:
        logger.error("--autofit-rows must be >= 0 (use 0 to disable auto-fit)")
        sys.exit(1)
    if args.autofit_max_width < 1:
        logger.error("--autofit-max-width must be >= 1")
        sys.exit(1)
    if args.max_validation_issues < 0:
        logger.error("--max-validation-issues must be >= 0 (use 0 for unlimited)")
        sys.exit(1)

    _, out_ext = os.path.splitext(args.output)
    if args.output_format in ("xlsx", "both", "all") and out_ext.lower() not in ("", ".xlsx"):
        logger.error(
            f"--output '{args.output}' has extension '{out_ext}'. "
            f"XLSX output requires a .xlsx filename (e.g. smartsheet.xlsx)."
        )
        sys.exit(1)

    # -- Fail-fast checks before any API work --------------------------------
    # 1. Parquet dependency
    _check_parquet_early(args.output_format, logger)

    # 2. Regex patterns -- compile and validate immediately
    def _compile_regex(pattern: str, flag_name: str) -> re.Pattern:
        try:
            return re.compile(pattern, re.IGNORECASE)
        except re.error as e:
            logger.error(f"Invalid regex for {flag_name}: {repr(pattern)} -- {e}")
            sys.exit(1)

    include_pat = _compile_regex(args.include_regex, "--include-regex") if args.include_regex else None
    exclude_pat = _compile_regex(args.exclude_regex, "--exclude-regex") if args.exclude_regex else None

    # Auth -- client created inside main, no module-level side effects
    token = os.environ.get("SMARTSHEET_API_TOKEN", "")
    if not token:
        logger.error("SMARTSHEET_API_TOKEN is not set. Add it to your .env file.")
        sys.exit(1)

    client = smartsheet.Smartsheet(token)
    client.errors_as_exceptions(True)
    try:
        with_retry(client.Users.get_current_user, logger=logger)
    except Exception as e:
        logger.error(f"Authentication failed: {e}")
        sys.exit(1)

    # Discovery
    records = discover_sheets(client, args.workspace_id, logger)

    # Apply name filters (patterns already compiled and validated above)
    if include_pat:
        before  = len(records)
        records = [r for r in records if include_pat.search(r.orig_name)]
        logger.info(f"--include-regex: kept {len(records)}/{before} sheet(s)")
    if exclude_pat:
        before  = len(records)
        records = [r for r in records if not exclude_pat.search(r.orig_name)]
        logger.info(f"--exclude-regex: kept {len(records)}/{before} sheet(s)")
    if args.max_sheets:
        records = records[:args.max_sheets]
        logger.info(f"--max-sheets: capped to {len(records)} sheet(s)")

    # Deterministic ordering: workspace -> sheet name -> sheet ID
    records.sort(key=lambda r: (r.workspace_name.lower(), r.orig_name.lower(), r.sheet_id))

    stats = ExportStats(total_found=len(records))

    # -- Incremental state ----------------------------------------------------
    state_path  = _state_path(args)
    state       = _load_state(state_path, logger)
    output_key  = os.path.abspath(args.output)
    run_ts      = datetime.now().strftime("%Y-%m-%d_%H%M%S")  # shared across all output formats

    since_dt = None
    if args.since:
        since_dt = _parse_since(args.since, state, output_key, logger)

    # Partition records: fetch from API vs restore from prior XLSX
    to_fetch  = records   # default: fetch all
    to_cache  = []        # records to restore from prior XLSX (incremental only)

    if since_dt is not None:
        prior_xlsx_exists = os.path.exists(args.output)
        if not prior_xlsx_exists:
            logger.warning(
                f"--since specified but no prior output found at '{args.output}'. "
                f"Running full export."
            )
        else:
            to_fetch = []
            to_cache = []
            for r in records:
                # Compare Smartsheet's own modified_at against the cutoff.
                # If modified_at is None (API didn't return it), always re-fetch -- safe default.
                if r.modified_at is not None and r.modified_at <= since_dt:
                    to_cache.append(r)
                else:
                    to_fetch.append(r)
            logger.info(
                f"Incremental mode (since {since_dt.isoformat()}): "
                f"{len(to_fetch)} sheet(s) to re-fetch, "
                f"{len(to_cache)} sheet(s) from cache."
            )

    # Dry run -- list sheets and exit
    if args.dry_run:
        logger.info("DRY RUN -- no files will be written")
        for r in records:
            status = "CACHE" if r in to_cache else "FETCH"
            logger.info(f"  [{r.workspace_name}]  {r.orig_name}  (id={r.sheet_id})  [{status}]")
        logger.info(f"Total after filters: {len(records)} sheet(s)")
        return

    # Parallel extraction (only sheets that need re-fetching)
    t_start  = time.time()
    fetched  = {}   # sheet_id -> (record, df | None, error | None)

    if to_fetch:
        logger.info(f"Fetching {len(to_fetch)} sheet(s) with up to {max(1, args.workers)} parallel workers...")

        def fetch_one(record):
            thread_client = smartsheet.Smartsheet(token)
            thread_client.errors_as_exceptions(True)
            return record, extract_sheet(thread_client, record, args.values, logger,
                                         row_metadata=args.row_metadata)

        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
            futures = {executor.submit(fetch_one, r): r for r in to_fetch}
            done    = 0
            for future in as_completed(futures):
                done  += 1
                record = futures[future]
                try:
                    _, df = future.result()
                    fetched[record.sheet_id] = (record, df, None)
                    if done == 1 or done % log_n == 0 or done == len(to_fetch):
                        logger.info(f"  [{done}/{len(to_fetch)}] {record.orig_name}")
                except Exception as e:
                    msg = f"SKIPPED '{record.orig_name}': {type(e).__name__}: {e}"
                    logger.error(msg)
                    stats.errors.append(msg)
                    stats.skipped_error += 1
                    fetched[record.sheet_id] = (record, None, msg)

    # Restore cached sheets from prior XLSX (incremental mode only)
    if to_cache:
        logger.info(f"Restoring {len(to_cache)} cached sheet(s) from {args.output}...")
        # Need the tab name mapping from the prior XLSX to look up by tab
        prior_tabs = {}
        if os.path.exists(args.output):
            try:
                prior_wb = load_workbook(args.output, read_only=True)
                prior_tabs = {name: name for name in prior_wb.sheetnames}
                prior_wb.close()
            except Exception as e:
                logger.warning(f"Could not inspect prior XLSX tabs: {e}. Will re-fetch cached sheets.")
                to_fetch.extend(to_cache)
                to_cache = []

        for record in to_cache:
            # Derive the tab name the prior run would have used (best-effort match)
            tab_name = sanitize_sheet_name(record.orig_name, fallback=f"Sheet_{record.sheet_id}")
            # Try exact match first, then sanitized
            matched_tab = prior_tabs.get(tab_name) or prior_tabs.get(record.orig_name)
            if matched_tab:
                df = _read_sheet_from_xlsx(args.output, matched_tab, logger)
            else:
                df = None
            if df is not None:
                fetched[record.sheet_id] = (record, df, None)
                logger.debug(f"  Cached: {record.orig_name}")
            else:
                # Cache miss -- fall back to fresh fetch
                logger.info(f"  Cache miss for '{record.orig_name}', re-fetching...")
                thread_client = smartsheet.Smartsheet(token)
                thread_client.errors_as_exceptions(True)
                try:
                    _, df = extract_sheet(thread_client, record, args.values, logger,
                                          row_metadata=args.row_metadata)
                    fetched[record.sheet_id] = (record, df, None)
                except Exception as e:
                    msg = f"SKIPPED '{record.orig_name}': {type(e).__name__}: {e}"
                    logger.error(msg)
                    stats.errors.append(msg)
                    stats.skipped_error += 1
                    fetched[record.sheet_id] = (record, None, msg)

    # Build manifest in original discovery order
    seen_tabs = {}
    manifest  = []
    for record in records:
        _, df, err = fetched.get(record.sheet_id, (record, None, "not fetched"))
        if err is not None:
            continue
        if df is None:
            stats.skipped_type += 1
            continue
        tab_name = unique_tab_name(record.orig_name, record.sheet_id, seen_tabs)
        manifest.append(FetchedSheet(record=record, df=df, tab_name=tab_name))

    index_tab = safe_index_name(seen_tabs)
    stem, ext = os.path.splitext(args.output)
    fmt       = args.output_format

    stats.exported = len(manifest)
    stats.elapsed  = time.time() - t_start

    # -- CSV output ----------------------------------------------------------
    if fmt in ("csv", "both", "all"):
        write_csv_output(manifest, stem, logger, run_ts=run_ts)
        if fmt == "csv":
            _log_summary(stats, logger, output=f"{stem}_csv/{run_ts}/")
            return

    # -- Parquet output ------------------------------------------------------
    if fmt in ("parquet", "all"):
        write_parquet_output(manifest, stem, logger, run_ts=run_ts)
        if fmt == "parquet":
            _log_summary(stats, logger, output=f"{stem}_parquet/{run_ts}/")
            return

    # -- XLSX output (default; also written for "both" and "all") -------------
    if fmt in ("xlsx", "both", "all"):
        tmp_path = f"{stem}_tmp_{uuid.uuid4().hex[:8]}{ext or '.xlsx'}"

        # Write to temp file (crash-safe: real file untouched until validated)
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name=index_tab, index=False)
            for fs in manifest:
                fs.df.to_excel(writer, sheet_name=fs.tab_name, index=False)

        # Post-write validation -- depth controlled by --validation-level
        logger.info(f"Validating output (level: {args.validation_level})...")
        wb_check = load_workbook(tmp_path, read_only=True)

        def row_digest(values) -> str:
            """Deterministic SHA-256 digest using unit-separator join -- avoids value-boundary collisions."""
            payload = "\x1f".join(_normalize(v) for v in values)
            return hashlib.sha256(payload.encode("utf-8")).hexdigest()

        for fs in manifest:
            if fs.tab_name not in wb_check.sheetnames:
                stats.validation_issues.append(f"Missing tab '{fs.record.orig_name}'")
                continue
            ws_chk  = wb_check[fs.tab_name]
            written = ws_chk.max_row - 1

            # All levels: row count
            if written != len(fs.df):
                msg = f"Row mismatch '{fs.record.orig_name}': expected {len(fs.df)}, got {written}"
                logger.warning(msg)
                stats.validation_issues.append(msg)

            # All levels: headers (cheap, catches high-value failures)
            written_headers  = [cell.value for cell in next(ws_chk.iter_rows(min_row=1, max_row=1))]
            expected_headers = list(fs.df.columns)
            if written_headers != expected_headers:
                msg = f"Header mismatch '{fs.record.orig_name}': expected {expected_headers}, got {written_headers}"
                logger.warning(msg)
                stats.validation_issues.append(msg)

            if args.validation_level == "basic" or len(fs.df) == 0:
                continue

            n = len(fs.df)

            if args.validation_level == "standard":
                # Sampled rows: quartiles + 3 seeded random -- reproducible per sheet
                rng        = random.Random(fs.record.sheet_id)
                candidates = sorted(set([
                    0, n // 4, n // 2, (3 * n) // 4, n - 1,
                    *rng.sample(range(n), min(3, n))
                ]))
                for row_idx in candidates:
                    src_digest  = row_digest(fs.df.iloc[row_idx].values)
                    xlsx_row    = next(ws_chk.iter_rows(
                        min_row=row_idx + 2, max_row=row_idx + 2, values_only=True
                    ))
                    if src_digest != row_digest(xlsx_row):
                        msg = f"Cell mismatch '{fs.record.orig_name}' row {row_idx + 1}: data differs after write"
                        logger.warning(msg)
                        stats.validation_issues.append(msg)
                        break   # one warning per sheet in standard mode

            else:
                # deep: every row -- report first N mismatches per sheet + total count
                # 0 = unlimited
                issue_cap  = args.max_validation_issues
                mismatch_count = 0
                for row_idx in range(n):
                    src_digest  = row_digest(fs.df.iloc[row_idx].values)
                    xlsx_row    = next(ws_chk.iter_rows(
                        min_row=row_idx + 2, max_row=row_idx + 2, values_only=True
                    ))
                    if src_digest != row_digest(xlsx_row):
                        mismatch_count += 1
                        if issue_cap == 0 or mismatch_count <= issue_cap:
                            msg = f"Cell mismatch '{fs.record.orig_name}' row {row_idx + 1}"
                            logger.warning(msg)
                            stats.validation_issues.append(msg)
                if mismatch_count > 0:
                    cap_note = (
                        f" (showing first {issue_cap})"
                        if issue_cap > 0 and mismatch_count > issue_cap else ""
                    )
                    logger.warning(
                        f"  -> {mismatch_count}/{n} row(s) with data divergence "
                        f"in '{fs.record.orig_name}'{cap_note}"
                    )
        wb_check.close()

        # Apply formatting
        autofit_rows = args.autofit_rows   # 0 = disabled, N = row cap
        max_width    = args.autofit_max_width
        fmt_mode     = args.format
        logger.info(f"Applying formatting (mode: {fmt_mode})...")
        wb = load_workbook(tmp_path)
        for fs in manifest:
            if fs.tab_name in wb.sheetnames:
                ws = wb[fs.tab_name]
                if fmt_mode == "pretty":
                    style_header_row(ws)
                    auto_fit_columns(ws, max_rows=autofit_rows, max_width=max_width)
                ws.freeze_panes = "A2"   # always applied -- zero cost, high usability value
        if not args.no_index:
            if index_tab in wb.sheetnames:
                del wb[index_tab]
            build_index_sheet(wb, manifest, index_tab,
                              fmt_mode=fmt_mode, autofit_rows=autofit_rows, max_width=max_width)
        elif index_tab in wb.sheetnames:
            del wb[index_tab]   # placeholder was written to ExcelWriter; remove it

        if not args.no_summary:
            summary_name = safe_summary_name(seen_tabs)
            skipped_name = safe_skipped_name(seen_tabs)
            build_summary_sheet(wb, stats, args, summary_name, skipped_name,
                                fmt_mode=fmt_mode, autofit_rows=autofit_rows, max_width=max_width)
        wb.save(tmp_path)

        # Archive prior output, then atomically promote temp file
        if os.path.exists(args.output):
            os.makedirs(ARCHIVE_DIR, exist_ok=True)
            ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
            stem2, _   = os.path.splitext(os.path.basename(args.output))
            old_backup = os.path.join(ARCHIVE_DIR, f"{stem2}_{ts}.xlsx")
            try:
                shutil.copy2(args.output, old_backup)
                logger.info(f"Archived previous output -> {old_backup}")
            except Exception as e:
                logger.warning(f"Archive failed ({e}). Continuing -- previous file will be overwritten.")

        try:
            os.replace(tmp_path, args.output)
        except Exception as e:
            logger.error(f"Failed to promote temp file: {e}")
            sys.exit(1)

        logger.info(f"XLSX written -> {args.output}")

        # Update incremental state sidecar -- just last_run timestamp.
        # Next --since last-run call will compare Smartsheet's own modified_at against this value.
        state[output_key] = {
            "last_run": datetime.now().isoformat(timespec="seconds"),
        }
        _save_state(state_path, state, logger)
        logger.debug(f"State sidecar updated -> {state_path}")

    _log_summary(stats, logger, output=args.output)


if __name__ == "__main__":
    main()
